"""
Generate the Verotel FlexPay test plan Excel.
Run from repo root: python docs/testing/generate_verotel_test_plan.py

Inherits the same structure and styling as generate_test_plan.py — re-uses
its helpers. Where the test names overlap with Stripe scenarios, the
Validation column is updated for Verotel specifics (saleID instead of
sub_xxx, custom1 round-trip, FlexPay postback events, etc.).
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "docs/testing/verotel-flexpay-test-plan.xlsx"

HEADER_BG  = "1A1A2E"
HEADER_FG  = "D4AF37"
CAT_BG     = "2D2D44"
CAT_FG     = "FFFFFF"
HIGH_BG    = "FFEBEE"
MED_BG     = "FFF8E1"
LOW_BG     = "E8F5E9"
STATUS_BG  = "F5F5F5"
ALT_ROW    = "FAFAFA"
WHITE      = "FFFFFF"
BORDER_COL = "CCCCCC"

thin = Side(style="thin", color=BORDER_COL)
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
PRIORITY_COLOURS = {"High": HIGH_BG, "Medium": MED_BG, "Low": LOW_BG}

COLUMNS = [
    ("ID", 9), ("Category", 20), ("Test Name", 34),
    ("Pre-conditions", 38), ("Test Steps", 52), ("Expected Result", 42),
    ("Validation / What to Check", 46), ("Priority", 10),
    ("Status", 12), ("Notes", 30),
]

TESTS = [

    # ── PROVIDER WIRING ───────────────────────────────────────────────────
    ("PR-01", "Provider wiring",
     "Provider config endpoint returns 'verotel'",
     "Backend up. PAYMENT_PROVIDER=verotel in env.",
     "1. GET /api/payments/config",
     "Returns {provider: 'verotel'}.",
     "Status 200. JSON body matches. Confirms config wiring.",
     "High"),

    ("PR-02", "Provider wiring",
     "Verotel provider fails closed if shop ID missing",
     "VEROTEL_SHOP_ID is blank in .env.",
     "1. Restart backend.\n2. Attempt POST /api/payments/checkout.",
     "Backend returns 503 / clear error. App doesn't expose half-broken checkout.",
     "Logs show ProviderNotConfiguredError. Endpoint responds with provider error.",
     "High"),

    ("PR-03", "Provider wiring",
     "Test mode flag respected",
     "VEROTEL_TEST_MODE=true.",
     "1. POST /api/payments/checkout.\n2. Inspect returned checkout URL.",
     "URL points at Verotel test/sandbox host (or test shopID), not live.",
     "URL contains test shopID or test-mode marker per Verotel docs.\nNo real charge possible.",
     "High"),

    # ── CHECKOUT URL CONSTRUCTION ─────────────────────────────────────────
    ("CK-01", "Checkout URL",
     "Essential monthly URL has all required FlexPay params",
     "Email-verified escort, free tier, no pending verification.",
     "1. POST /api/payments/checkout {tier:'essential', billing:'monthly'}.\n2. Parse the returned URL.",
     "URL contains version, type=subscription, shopID, priceAmount, priceCurrency=GBP, period=P1M, email, custom1, signature.",
     "Inspect URL query string — every required param present and correct.",
     "High"),

    ("CK-02", "Checkout URL",
     "Annual billing produces period=P1Y",
     "Free-tier verified escort.",
     "1. POST checkout with billing='annual'.",
     "URL period param equals P1Y. priceAmount matches platform_config annual amount.",
     "Query string inspection.",
     "High"),

    ("CK-03", "Checkout URL",
     "Signature is valid and matches our local computation",
     "Any valid checkout creation.",
     "1. Create a checkout session.\n2. Extract signature from URL.\n3. Recompute locally using VEROTEL_SIGNATURE_KEY and sorted params.",
     "Computed and returned signatures match exactly.",
     "Verify our HMAC-SHA256 implementation matches Verotel's expectation.",
     "High"),

    ("CK-04", "Checkout URL",
     "Discount applies to first charge only, not rebill",
     "FM-xxx code with 50% off active.",
     "1. POST checkout with discount_code='FM-xxx'.\n2. Inspect URL.",
     "priceAmount is ~50% of normal first charge. subscriptionPriceAmount is the full price for rebills.",
     "Reduced priceAmount + nextChargeOn + subscriptionPriceAmount all present.",
     "High"),

    ("CK-05", "Checkout URL",
     "Metadata round-trips via custom1",
     "Any checkout.",
     "1. Create checkout.\n2. Parse custom1 JSON from URL.",
     "custom1 contains escort_id, tier, billing as a JSON object.",
     "JSON parse of custom1 returns the metadata we passed.",
     "High"),

    ("CK-06", "Checkout URL",
     "Blue Tick combines setup fee + monthly into first charge",
     "Essential-tier escort with no Blue Tick.",
     "1. POST /api/payments/blue-tick-checkout.",
     "priceAmount = setup_pence + monthly_pence (e.g. £13.99 = £10 + £3.99).\nsubscriptionPriceAmount = monthly only (£3.99).",
     "Inspect URL params. Verify decimal math.",
     "High"),

    # ── HOSTED CHECKOUT FLOW ──────────────────────────────────────────────
    ("HC-01", "Hosted checkout",
     "User redirected from app to Verotel checkout page",
     "Valid checkout URL returned by backend.",
     "1. Frontend calls createCheckout().\n2. window.location.href = url.",
     "Browser navigates to secure.verotel.com page. Verotel logo + correct product name shown.",
     "Visual check on Verotel hosted page. URL bar shows verotel.com domain.",
     "High"),

    ("HC-02", "Hosted checkout",
     "Test card 4111 1111 1111 1111 (Verotel test mode) succeeds",
     "VEROTEL_TEST_MODE=true. On Verotel test checkout page.",
     "1. Enter 4111 1111 1111 1111, any future expiry, any CVV.\n2. Submit.",
     "Verotel processes test charge. Redirects to our success_url.",
     "URL after redirect matches success_url. Verotel test dashboard shows test sale.",
     "High"),

    ("HC-03", "Hosted checkout",
     "Successful payment redirects to success_url with Verotel-signed params",
     "Just completed test payment.",
     "1. Land on /dashboard/verify?payment=success&saleID=...&signature=...",
     "URL contains saleID + signature. Our success page handles the redirect.",
     "Browser is on /dashboard/verify page. Frontend doesn't crash.",
     "High"),

    ("HC-04", "Hosted checkout",
     "Declined card returns to cancel_url",
     "On Verotel test checkout. Use a declined test card (per Verotel docs).",
     "1. Enter declined card.\n2. Submit.",
     "Verotel shows decline error. Optional 'try again'. User redirected to cancel_url after giving up.",
     "Land on /dashboard/subscription?payment=cancelled. No DB changes.",
     "High"),

    # ── POSTBACK / WEBHOOK ────────────────────────────────────────────────
    ("WH-01", "Webhook",
     "initial postback creates subscription in DB",
     "User completed a test checkout.",
     "1. Verotel sends 'initial' postback to /api/webhooks/verotel.\n2. Check subscriptions table.",
     "New row created with status='active', tier from metadata, psp_subscription_id = saleID.",
     "DB row exists. escort.subscription_tier updated. escort.psp_subscription_id set.",
     "High"),

    ("WH-02", "Webhook",
     "Invalid signature is rejected with 400",
     "Webhook secret known. Send a postback with a tampered amount.",
     "1. POST /api/webhooks/verotel with valid-shape body but signature recomputed from tampered fields.",
     "Endpoint returns 400. No DB mutations.",
     "Webhook events table has NO entry for the bad request. Subscriptions table unchanged.",
     "High"),

    ("WH-03", "Webhook",
     "Missing signature is rejected with 400",
     "Send postback with `signature` field omitted entirely.",
     "1. POST /api/webhooks/verotel.",
     "Endpoint returns 400 'Invalid webhook signature'.",
     "InvalidWebhookSignatureError logged. No DB writes.",
     "High"),

    ("WH-04", "Webhook",
     "Duplicate delivery is idempotent",
     "An 'initial' postback has already been processed.",
     "1. Resend the same postback (same saleID, event, txn).",
     "Second delivery returns {received: true, duplicate: true}. DB unchanged.",
     "webhook_events table has 1 row for the dedup key. No duplicate subscriptions row.",
     "High"),

    ("WH-05", "Webhook",
     "rebill event updates current_period_end",
     "Active subscription. ~30 days after initial.",
     "1. Verotel sends 'rebill' postback.\n2. Check subscriptions row.",
     "current_period_start / current_period_end advance by one month. Status remains 'active'.",
     "DB period fields updated. subscription_expires_at on escort updated.",
     "High"),

    ("WH-06", "Webhook",
     "rebill applies pending downgrade",
     "Active premium sub with pending_tier='essential'.",
     "1. Wait for / simulate next 'rebill' postback.",
     "tier and escort.subscription_tier flip to 'essential'. pending_tier cleared.",
     "DB updates correctly. Photo-limit check fires (paused if over limit).",
     "High"),

    ("WH-07", "Webhook",
     "cancel event marks status='cancelling'",
     "Active subscription.",
     "1. User clicks Cancel.\n2. Backend calls provider.cancel_subscription.\n3. Verotel sends 'cancel' postback.",
     "DB row status='cancelling'. Access continues until period end.",
     "subscriptions.status='cancelling'. escort.subscription_tier still set.",
     "High"),

    ("WH-08", "Webhook",
     "expiry event ends access",
     "Sub in 'cancelling' state. Period has passed.",
     "1. Verotel sends 'expiry' postback.",
     "status='cancelled'. escort.subscription_tier='free'. psp_subscription_id cleared.",
     "DB. UI shows Free tier on next page load.",
     "High"),

    ("WH-09", "Webhook",
     "uncancel event reactivates",
     "Sub in 'cancelling' state. Admin un-cancelled in Verotel Control Center.",
     "1. Verotel sends 'uncancel' postback.",
     "status returns to 'active'.",
     "DB row updated. UI shows Active.",
     "Medium"),

    ("WH-10", "Webhook",
     "credit (refund) postback recorded",
     "Refund issued in Verotel Control Center.",
     "1. Verotel sends 'credit' postback.",
     "subscriptions.cancelled_at set to mark the refund event. Admin should review whether to revoke access.",
     "DB shows refund timestamp. Admin alerted.",
     "Medium"),

    ("WH-11", "Webhook",
     "chargeback instantly revokes access",
     "Customer disputed a charge.",
     "1. Verotel sends 'chargeback' postback.",
     "subscription cancelled immediately. escort.subscription_tier='free'. escort.is_approved=false.",
     "DB. Escort no longer appears in search.",
     "High"),

    ("WH-12", "Webhook",
     "Unknown event type is logged but returns 200",
     "Verotel sends an event we don't handle.",
     "1. POST event with type='something_new'.",
     "Returns 200 (so Verotel doesn't retry forever). Log message captured.",
     "Server logs show '[verotel webhook] Unknown event type'. No DB writes.",
     "Medium"),

    ("WH-13", "Webhook",
     "Webhook handles both POST and GET",
     "Verotel may send GETs in some scenarios.",
     "1. Test the same postback as both POST body and GET query params.",
     "Both succeed identically.",
     "Same DB outcome from both methods.",
     "Medium"),

    # ── SUBSCRIPTION LIFECYCLE ────────────────────────────────────────────
    ("SL-01", "Subscription lifecycle",
     "Cancel via /payments/cancel calls Verotel API",
     "Active subscription.",
     "1. POST /api/payments/cancel.",
     "Verotel cancel-subscription API hit with correct shopID + saleID + signature. Local sub marked 'cancelling'.",
     "Inspect outgoing request via mock or staging server. DB updated.",
     "High"),

    ("SL-02", "Subscription lifecycle",
     "Cancel of already-cancelled sub is idempotent",
     "Sub already cancelled in Verotel.",
     "1. Call /payments/cancel again.",
     "Returns success. Verotel 404 treated as success. No errors surfaced to user.",
     "Backend doesn't throw 502 on idempotent cancel.",
     "Medium"),

    ("SL-03", "Subscription lifecycle",
     "Cancel without active sub returns 400",
     "Free-tier escort.",
     "1. POST /payments/cancel.",
     "Returns 400 'No active subscription'.",
     "API responds 400. Clear UI message.",
     "Medium"),

    ("SL-04", "Subscription lifecycle",
     "Upgrade triggers cancel + new checkout, NOT in-place modify",
     "Essential subscriber wants Premium.",
     "1. POST /api/payments/upgrade-tier {tier:'premium'}.",
     "Old sub cancelled. Response returns a NEW checkout URL for premium.\nFrontend redirects user to it.",
     "Verotel API cancel was called. Response payload contains {url}. User redirected to Verotel page.",
     "High"),

    ("SL-05", "Subscription lifecycle",
     "Reactivation message instructs user to re-subscribe",
     "Cancelling subscription, before period end.",
     "1. User clicks Reactivate in MySubscriptionsPage.",
     "Toast: 'To reactivate, start a new subscription'.",
     "No backend call (Verotel can't reactivate via API). User pointed to Subscription page.",
     "Medium"),

    # ── DISCOUNT CODES + FOUNDING ─────────────────────────────────────────
    ("DC-01", "Discount codes",
     "FM- code applies 50% off first charge",
     "Active FM-xxx code with 50%, valid for all tiers.",
     "1. POST /payments/checkout with discount_code='FM-xxx'.\n2. Inspect URL.",
     "First-charge priceAmount = 50% of full price. Subscription rebill price = full.",
     "URL math correct.",
     "High"),

    ("DC-02", "Discount codes",
     "FM- code redemption counter increments",
     "DiscountCode current_redemptions=0.",
     "1. Complete checkout with the code.",
     "current_redemptions=1 after successful checkout creation.",
     "DB check.",
     "High"),

    ("DC-03", "Discount codes",
     "FM- code grants founding member status",
     "Escort uses FM-xxx code at checkout (not at register).",
     "1. POST checkout with FM- code.",
     "escort.is_founding_member=true. founding_offer_signups incremented. Prospect linked.",
     "DB: escort + platform_config + outreach_prospects all updated atomically.",
     "High"),

    ("DC-04", "Discount codes",
     "Tier-restricted code rejects wrong tier",
     "Code with applicable_tiers=['premium'].",
     "1. Try to apply on Essential checkout.",
     "Returns 400 with clear message.",
     "No URL returned. No redemption incremented.",
     "Medium"),

    ("DC-05", "Discount codes",
     "Referral code grants 50% off 3 months",
     "Valid referral_code on another escort.",
     "1. POST checkout with referral_code.",
     "first-charge discounted 50%. escort.referred_by_code set.",
     "URL math correct. DB updated.",
     "Medium"),

    # ── BLUE TICK ─────────────────────────────────────────────────────────
    ("BT-01", "Blue Tick",
     "Blue Tick checkout blocked for Premium/Elite",
     "Premium subscriber.",
     "1. POST /api/payments/blue-tick-checkout.",
     "Returns 400 with helpful message.",
     "No URL. User sees explanatory error.",
     "High"),

    ("BT-02", "Blue Tick",
     "Blue Tick checkout creates pending L3 verification on success",
     "Essential subscriber. Pay Blue Tick via test card.",
     "1. Complete checkout flow.\n2. Verotel sends 'initial' postback.",
     "Verification row created (level=3, status='pending'). Admin email sent.",
     "DB row. Admin email logged in dev logs.",
     "High"),

    ("BT-03", "Blue Tick",
     "Cannot purchase Blue Tick twice",
     "Already has psp_blue_tick_subscription_id.",
     "1. POST blue-tick-checkout again.",
     "Returns 400 'already have Blue Tick'.",
     "No second checkout URL.",
     "Medium"),

    # ── EDGE CASES ────────────────────────────────────────────────────────
    ("EC-01", "Edge cases",
     "Checkout during pending verification returns 409",
     "Pending L2 verification exists.",
     "1. POST checkout.",
     "Returns 409 with explanation.",
     "No checkout URL.",
     "High"),

    ("EC-02", "Edge cases",
     "Webhook missing escort_id in metadata falls back to saleID lookup",
     "Postback has empty custom1.",
     "1. Send postback with custom1=''.",
     "Webhook looks up by psp_subscription_id and finds the escort by that field.",
     "Subscription row updated correctly even without metadata.",
     "Medium"),

    ("EC-03", "Edge cases",
     "Webhook completely orphaned (no escort matches) is silently dropped",
     "Postback references a saleID we never recorded.",
     "1. Send postback.",
     "Returns 200. Log entry shows could-not-find-escort.",
     "No DB writes. No 500.",
     "Medium"),

    ("EC-04", "Edge cases",
     "Concurrent cancel + rebill race doesn't corrupt state",
     "User cancels at the exact moment a rebill postback arrives.",
     "1. Send 'rebill' and 'cancel' postbacks within ms of each other.",
     "Final state is consistent — either cancelling+renewed (period extended, will cancel next cycle) or just cancelling. Never both 'active' and 'cancelling' on same row.",
     "DB consistency check. No exceptions raised.",
     "Low"),

    # ── PROD READINESS ────────────────────────────────────────────────────
    ("PD-01", "Production readiness",
     "VEROTEL_TEST_MODE=false uses live endpoints",
     "Flip env var to false. Restart backend.",
     "1. Create checkout.",
     "URL points at live Verotel checkout host. WARNING displayed on staging if any real cards used.",
     "URL contains live shopID + live host.",
     "High"),

    ("PD-02", "Production readiness",
     "Webhook URL configured in Verotel Control Center matches our route",
     "Verotel Control Center config check.",
     "1. Log into Verotel.\n2. Check postback URL is https://api.bluechips.live/api/webhooks/verotel.",
     "URL exact match. HTTPS only.",
     "Visual check.",
     "High"),

    ("PD-03", "Production readiness",
     "Signature key + API user/pass match between Verotel and .env",
     "Production .env populated.",
     "1. Run a test checkout against live Verotel.",
     "Verotel accepts signature. Cancel API call succeeds.",
     "End-to-end live test.",
     "High"),

    ("PD-04", "Production readiness",
     "All FM- codes in prod open to all tiers",
     "Prod DB.",
     "1. SELECT code, applicable_tiers FROM discount_codes WHERE code LIKE 'FM-%';",
     "All show applicable_tiers='{}'.",
     "DB query result.",
     "High"),
]


def _cell_style(ws, row, col, value, bg=None, fg="000000", bold=False,
                wrap=True, align="left", border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(color=fg, bold=bold, size=10, name="Calibri")
    cell.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal=align)
    if border:
        cell.border = THIN_BORDER
    return cell


def _build_test_sheet(ws):
    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "Bluechips London — Verotel FlexPay Test Plan"
    t.fill = PatternFill("solid", fgColor=HEADER_BG)
    t.font = Font(color=HEADER_FG, bold=True, size=14, name="Calibri")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    s = ws["A2"]
    s.value = ("Sandbox test cards: Verotel docs provide test PANs after approval — "
               "DO NOT use Stripe test cards (won't work). Verify TEST MODE banner.")
    s.fill = PatternFill("solid", fgColor="2D2D44")
    s.font = Font(color="AAAAAA", italic=True, size=9, name="Calibri")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=3, column=col_idx, value=col_name)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.font = Font(color=HEADER_FG, bold=True, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "D4"

    current_cat = None
    row = 4
    for test in TESTS:
        tid, cat, name, pre, steps, expected, validation, priority = test
        if cat != current_cat:
            current_cat = cat
            ws.merge_cells(f"A{row}:J{row}")
            cc = ws[f"A{row}"]
            cc.value = f"  {cat.upper()}"
            cc.fill = PatternFill("solid", fgColor=CAT_BG)
            cc.font = Font(color=CAT_FG, bold=True, size=10, name="Calibri")
            cc.alignment = Alignment(vertical="center")
            cc.border = THIN_BORDER
            ws.row_dimensions[row].height = 18
            row += 1

        row_bg = ALT_ROW if row % 2 == 0 else WHITE
        pri_bg = PRIORITY_COLOURS.get(priority, WHITE)
        values = [tid, cat, name, pre, steps, expected, validation, priority, "Not Run", ""]
        for col_idx, val in enumerate(values, start=1):
            bg = pri_bg if col_idx == 8 else (STATUS_BG if col_idx == 9 else row_bg)
            fg = "555555" if col_idx == 1 else "000000"
            _cell_style(ws, row, col_idx, val, bg=bg, fg=fg, bold=(col_idx == 1))
        ws.row_dimensions[row].height = max(60, max(len(str(v)) for v in [steps, expected, validation]) // 3)
        row += 1

    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"Not Run,Pass,Fail,Blocked,Skip"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.sqref = f"I4:I{row}"


def _build_summary_sheet(wb):
    ws = wb.create_sheet("Summary")
    for col, w in zip("ABCDEF", (28, 12, 12, 12, 12, 24)):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:F1")
    t = ws["A1"]; t.value = "Verotel Test Plan Summary"
    t.fill = PatternFill("solid", fgColor=HEADER_BG)
    t.font = Font(color=HEADER_FG, bold=True, size=13, name="Calibri")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(["Category", "Total", "High", "Medium", "Low", "Notes"], 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.font = Font(color=HEADER_FG, bold=True, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER

    from collections import defaultdict
    counts = defaultdict(lambda: {"total": 0, "High": 0, "Medium": 0, "Low": 0})
    for t in TESTS:
        cat, pri = t[1], t[-1]
        counts[cat]["total"] += 1
        counts[cat][pri] += 1

    row = 3
    totals = {"total": 0, "High": 0, "Medium": 0, "Low": 0}
    for cat, c in counts.items():
        for col, val in enumerate([cat, c["total"], c["High"], c["Medium"], c["Low"], ""], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")
            cell.border = THIN_BORDER
        for k in totals:
            totals[k] += c[k]
        row += 1

    for col, val in enumerate(["TOTAL", totals["total"], totals["High"], totals["Medium"], totals["Low"], ""], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=CAT_BG)
        cell.font = Font(color=CAT_FG, bold=True, size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left")
        cell.border = THIN_BORDER


if __name__ == "__main__":
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    _build_test_sheet(ws)
    _build_summary_sheet(wb)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")
    print(f"Total tests: {len(TESTS)} ({sum(1 for t in TESTS if t[-1] == 'High')} High priority)")
