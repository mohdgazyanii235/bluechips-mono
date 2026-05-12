"""
Generate payment & subscription test plan Excel workbook.
Run from repo root: python docs/testing/generate_test_plan.py
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

OUTPUT = "docs/testing/payment-subscription-test-plan.xlsx"

# ── Colour palette ───────────────────────────────────────────────────────────
HEADER_BG   = "1A1A2E"   # dark navy
HEADER_FG   = "D4AF37"   # gold
CAT_BG      = "2D2D44"   # dark section header
CAT_FG      = "FFFFFF"
HIGH_BG     = "FFEBEE"   # red tint
MED_BG      = "FFF8E1"   # amber tint
LOW_BG      = "E8F5E9"   # green tint
STATUS_BG   = "F5F5F5"   # light grey for status cells
ALT_ROW     = "FAFAFA"
WHITE       = "FFFFFF"
BORDER_COL  = "CCCCCC"

thin = Side(style="thin", color=BORDER_COL)
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

PRIORITY_COLOURS = {"High": HIGH_BG, "Medium": MED_BG, "Low": LOW_BG}

# ── Column definitions ───────────────────────────────────────────────────────
COLUMNS = [
    ("ID",           9),
    ("Category",     18),
    ("Test Name",    32),
    ("Pre-conditions", 38),
    ("Test Steps",   52),
    ("Expected Result", 42),
    ("Validation / What to Check", 42),
    ("Priority",     10),
    ("Status",       12),
    ("Notes",        30),
]

# ── Test cases ───────────────────────────────────────────────────────────────
# Each entry: (id, category, name, preconditions, steps, expected, validation, priority)
TESTS = [

    # ── FOUNDING MEMBER FLOW ─────────────────────────────────────────────
    ("FM-01", "Founding Member",
     "Public counter displays correctly",
     "Founding offer is active in admin config with a limit set",
     "1. Visit /join page.\n2. Check the founding member counter widget.",
     "Counter shows correct remaining spots (e.g. '23 of 50 remaining').",
     "API: GET /api/founding/status → active=true, remaining = limit - signups.\nUI counter matches API.",
     "High"),

    ("FM-02", "Founding Member",
     "Register via founding member invite link",
     "A valid FM-XXXX code exists and has remaining uses",
     "1. Open a founding member invite link (/?code=FM-XXXX).\n2. Click 'Join'.\n3. Complete registration form.\n4. Submit.",
     "Registration succeeds. User is flagged as founding member. signup_discount_code_id set on escort record.",
     "DB: escorts.is_founding_member = true, escorts.signup_discount_code_id = ID of FM code.\nFounding offer signups incremented by 1.",
     "High"),

    ("FM-03", "Founding Member",
     "Founding offer auto-closes at limit",
     "Founding offer limit is set to N; signups = N - 1",
     "1. Register the last available founding member slot.\n2. Try to register with the same FM code again (new account).",
     "Second signup attempt shows 'This offer has ended' or similar. Counter shows 0 remaining.",
     "API: GET /api/founding/status → active=false (or remaining=0).\nSecond registration attempt: 400/409 error returned.\nplatform_config.founding_offer_signups = limit.",
     "High"),

    ("FM-04", "Founding Member",
     "Founding discount auto-applies at checkout",
     "User registered with FM code (signup_discount_code_id set). User is email-verified.",
     "1. Log in as founding member.\n2. Go to subscription page.\n3. Select any paid tier, click Subscribe.",
     "Checkout session is created with the founding member coupon applied. Stripe checkout page shows the discounted price.",
     "Stripe Dashboard: checkout session has a coupon/discount attached.\nDB: discount_codes.current_redemptions incremented by 1 after checkout.",
     "High"),

    ("FM-05", "Founding Member",
     "FM discount applied only once per user",
     "Founding member has already subscribed once using their FM code",
     "1. Log in as founding member who already used their FM code.\n2. Cancel subscription.\n3. Subscribe again.",
     "On second subscription, no founding member discount is applied automatically. Full price shown.",
     "DB: discount_codes.current_redemptions = max_redemptions for that user's code.\nStripe checkout: no coupon applied.",
     "Medium"),

    ("FM-06", "Founding Member",
     "FM discount respects applicable tier restriction",
     "FM code has applicable_tiers = ['essential'] only",
     "1. Log in as founding member.\n2. Attempt checkout for Premium tier.",
     "Checkout fails with error: 'This code is only valid for: Essential'.",
     "API returns 400 with tier restriction message.\nDiscount code current_redemptions NOT incremented.",
     "Medium"),

    # ── DISCOUNT CODES ────────────────────────────────────────────────────
    ("DC-01", "Discount Codes",
     "Valid % discount on monthly subscription",
     "Active discount code with 20% off for 3 months exists. User has no existing subscription.",
     "1. Go to checkout.\n2. Enter discount code in the code field.\n3. Select Essential monthly.\n4. Complete checkout.",
     "Stripe checkout shows 20% discount applied. First 3 invoices are 20% off. Month 4 reverts to full price.",
     "Stripe Dashboard: subscription has coupon with 20% off × 3 months.\nDB: current_redemptions incremented.\nInvoice history shows correct amounts.",
     "High"),

    ("DC-02", "Discount Codes",
     "Valid % discount on annual subscription (pro-rata)",
     "Active discount code with 20% off for 3 months. User selects annual billing.",
     "1. Go to checkout.\n2. Enter discount code.\n3. Select Essential annual.\n4. Complete checkout.",
     "Annual checkout applies a one-time amount_off coupon equivalent to (annual_price/12) × 20% × 3 months.",
     "Stripe Dashboard: coupon is 'once' duration with correct GBP amount_off calculated.\nStripe checkout page shows discounted annual price.",
     "High"),

    ("DC-03", "Discount Codes",
     "Invalid / inactive code rejected",
     "No discount code with that name, or code is_active = false",
     "1. Go to checkout.\n2. Enter 'FAKECODE123' in discount field.\n3. Submit.",
     "API returns 400: 'Discount code is invalid or no longer active'.",
     "Error message displayed in UI.\nNo Stripe session created.\nDB current_redemptions unchanged.",
     "High"),

    ("DC-04", "Discount Codes",
     "Code at usage limit rejected",
     "Discount code has max_redemptions=10, current_redemptions=10",
     "1. Go to checkout.\n2. Enter the exhausted code.\n3. Submit.",
     "API returns 400: 'This code has reached its maximum number of uses'.",
     "Error message displayed in UI.\nNo Stripe session created.",
     "High"),

    ("DC-05", "Discount Codes",
     "Tier-restricted code rejects wrong tier",
     "Code applicable_tiers = ['premium', 'elite'] only",
     "1. Go to checkout.\n2. Enter restricted code.\n3. Select Essential tier.\n4. Submit.",
     "API returns 400: 'This code is only valid for: Premium, Elite'.",
     "Error message shown. Redemption count unchanged.",
     "Medium"),

    ("DC-06", "Discount Codes",
     "Discount visible in invoice history",
     "User paid using a discount code",
     "1. Log in.\n2. Go to My Subscriptions / Invoices page.",
     "Invoice history shows the discounted amount actually charged, not the full price.",
     "GET /api/payments/invoices → invoice amount_paid matches discounted amount.\nInvoice PDF (from Stripe) shows coupon applied.",
     "Medium"),

    # ── NEW SUBSCRIPTIONS ─────────────────────────────────────────────────
    ("NS-01", "New Subscription",
     "Essential monthly checkout session created",
     "Email-verified escort with no existing subscription. Stripe keys configured.",
     "1. Log in.\n2. Subscription page → select Essential Monthly.\n3. Click Subscribe.",
     "Redirected to Stripe Checkout URL. Page shows Essential plan at £24.99/mo.",
     "API: POST /payments/checkout returns {url: 'https://checkout.stripe.com/...'}.\nStripe Dashboard: checkout session in 'open' state.",
     "High"),

    ("NS-02", "New Subscription",
     "Essential monthly payment completes",
     "Checkout session created for Essential monthly",
     "1. On Stripe Checkout, use test card 4242 4242 4242 4242.\n2. Complete payment.\n3. Return to site.",
     "Redirected to /dashboard/verify?payment=success. Tier updated to 'essential'. Subscription record created.",
     "DB: escorts.subscription_tier = 'essential', escorts.stripe_subscription_id = sub ID.\nDB: subscriptions table has new row with tier='essential', status='active'.\nUI: Dashboard shows 'Essential' plan.",
     "High"),

    ("NS-03", "New Subscription",
     "Essential annual checkout + payment",
     "Email-verified escort with no subscription. Annual price ID configured.",
     "1. Select Essential Annual.\n2. Complete checkout with test card.",
     "Subscription created at annual price. Period end is ~1 year from billing anchor.",
     "DB: subscription_tier='essential', period_end ~= Jan 1 next year.\nStripe: subscription interval = 'year'.",
     "High"),

    ("NS-04", "New Subscription",
     "Premium monthly checkout + payment",
     "Email-verified escort. No existing subscription.",
     "1. Select Premium Monthly.\n2. Complete checkout.",
     "Tier updated to 'premium'. Photo limit becomes 50.",
     "DB: subscription_tier='premium'.\nUI: photo upload shows 50 slots available.",
     "High"),

    ("NS-05", "New Subscription",
     "Elite monthly checkout + payment",
     "Email-verified escort. No existing subscription.",
     "1. Select Elite Monthly.\n2. Complete checkout.",
     "Tier updated to 'elite'.",
     "DB: subscription_tier='elite'.\nUI: Dashboard shows Elite plan badge.",
     "High"),

    ("NS-06", "New Subscription",
     "Billing anchor set to 1st of next month",
     "User starts a monthly subscription mid-month",
     "1. Subscribe on any day that is not the 1st.\n2. Complete payment.",
     "First invoice is a partial-month proration. Subsequent invoices bill on the 1st.",
     "Stripe Dashboard: subscription billing_cycle_anchor = 1st of next month (Unix timestamp).\nFirst invoice amount < full monthly price (prorated).",
     "Medium"),

    ("NS-07", "New Subscription",
     "Duplicate checkout - existing active sub replaced",
     "Escort has active Essential subscription",
     "1. Somehow reach the checkout flow again (e.g. direct URL).\n2. Complete payment for Essential.",
     "Old subscription record marked 'replaced'. New record created as active.",
     "DB: old subscriptions row status='replaced'.\nOnly one 'active' subscription row per escort for non-blue-tick.",
     "Medium"),

    ("NS-08", "New Subscription",
     "Annual billing unavailable returns 503",
     "Annual Stripe price ID is not configured (blank in platform_config)",
     "1. Select any annual plan.\n2. Click Subscribe.",
     "API returns 503: 'Annual billing is not yet available. Please select monthly billing.'",
     "Error message displayed. No Stripe session created.",
     "Low"),

    # ── UPGRADES ─────────────────────────────────────────────────────────
    ("UP-01", "Upgrade",
     "Upgrade preview (Essential → Premium)",
     "Escort is on Essential monthly. Billing is mid-month.",
     "1. Go to Subscription page.\n2. Select Premium.\n3. View upgrade preview/confirmation modal.",
     "Modal shows: pro-rata charge for remaining days in month, then £49.99/mo from 1st.",
     "GET /api/payments/upgrade-preview?tier=premium&billing=monthly → charge_now_pence = correct pro-rata.\nUI modal shows correct GBP amounts.",
     "High"),

    ("UP-02", "Upgrade",
     "Essential → Premium upgrade (immediate charge)",
     "Escort is on Essential monthly with a valid payment method on file.",
     "1. Confirm upgrade from Essential to Premium.\n2. Approve.",
     "Pro-rata invoice created and charged immediately. Subscription price updated. Tier = 'premium' in DB.",
     "Stripe Dashboard: invoice created for pro-rata amount, status='paid'.\nStripe: subscription item now points to Premium price ID.\nDB: escorts.subscription_tier='premium'.\nUI: Dashboard shows Premium.",
     "High"),

    ("UP-03", "Upgrade",
     "Essential → Elite upgrade (immediate charge)",
     "Escort is on Essential monthly.",
     "1. Select Elite → Confirm upgrade.",
     "Higher pro-rata invoice charged. Tier becomes 'elite'.",
     "Stripe invoice for Elite-Essential difference × remaining_days/total_days.\nDB: subscription_tier='elite'.",
     "High"),

    ("UP-04", "Upgrade",
     "Premium → Elite upgrade",
     "Escort is on Premium monthly.",
     "1. Select Elite → Confirm upgrade.",
     "Smaller pro-rata invoice (Elite − Premium difference). Tier = 'elite'.",
     "Stripe invoice for correct amount.\nDB: subscription_tier='elite'.",
     "High"),

    ("UP-05", "Upgrade",
     "Upgrade uses same Stripe subscription (not new)",
     "Escort is on Essential. Performs upgrade to Premium.",
     "1. Upgrade to Premium.\n2. Check Stripe Dashboard.",
     "The existing stripe_subscription_id is updated, NOT a new subscription created.",
     "DB: escorts.stripe_subscription_id is the SAME value before and after upgrade.\nStripe: subscription items updated.",
     "High"),

    ("UP-06", "Upgrade",
     "Auto-grant Blue Tick on upgrade to Premium/Elite (identity already verified)",
     "Escort has verification_level=2 (identity verified) and is on Essential.",
     "1. Upgrade to Premium.",
     "blue_tick_active becomes True automatically. No extra payment needed.",
     "DB: escorts.blue_tick_active=true after upgrade.\nUI: Blue Tick badge appears on profile.",
     "High"),

    ("UP-07", "Upgrade",
     "Upgrade fails if payment fails",
     "Escort on Essential. Use a test card that declines (4000 0000 0000 0002).",
     "1. Attempt upgrade to Premium.\n2. Payment method on file is a decline card.",
     "API returns 402: 'Payment failed: ...' message. Subscription price NOT changed. Tier stays 'essential'.",
     "DB: subscription_tier still 'essential'.\nStripe: no successful invoice for upgrade.\nVoided invoice visible in Stripe (if invoice was created).",
     "High"),

    ("UP-08", "Upgrade",
     "Upgrade to same plan rejected",
     "Escort is on Premium monthly.",
     "1. Call upgrade endpoint with tier=premium, billing=monthly.",
     "API returns 400: 'You are already on this plan with this billing period'.",
     "No Stripe calls made. Clear error message in UI.",
     "Medium"),

    # ── DOWNGRADES ────────────────────────────────────────────────────────
    ("DG-01", "Downgrade",
     "Downgrade preview (Premium → Essential)",
     "Escort is on Premium monthly.",
     "1. Go to Subscription page.\n2. Select Essential.\n3. View confirmation modal.",
     "Modal shows: no charge today, downgrade effective from next billing date.",
     "GET /api/payments/upgrade-preview?tier=essential → type='downgrade', charge_now_pence=0, effective_date shown.",
     "High"),

    ("DG-02", "Downgrade",
     "Downgrade scheduled — no immediate charge",
     "Escort is on Elite monthly.",
     "1. Confirm downgrade to Essential.",
     "No immediate invoice. Tier stays Elite until end of billing period. pending_tier='essential' in DB.",
     "Stripe Dashboard: subscription still shows Elite price.\nDB: subscriptions.pending_tier='essential'.\nDB: escorts.subscription_tier still 'elite'.",
     "High"),

    ("DG-03", "Downgrade",
     "Downgrade takes effect at billing renewal",
     "Escort downgraded (pending_tier set). Billing period ends.",
     "1. Simulate/wait for subscription.updated webhook from Stripe with new period start.\n2. Check DB + UI.",
     "Tier updates to downgraded tier. pending_tier cleared.",
     "DB: escorts.subscription_tier = downgraded tier.\nDB: subscriptions.pending_tier = null.\nUI: Dashboard shows new lower tier.",
     "High"),

    ("DG-04", "Downgrade",
     "Photo warning email sent when downgrading over photo limit",
     "Escort on Premium has 15 photos. Downgrading to Essential (limit: 8).",
     "1. Confirm downgrade from Premium to Essential.",
     "Email sent to escort warning they have 7 excess photos and must remove them before billing date.",
     "Email received with photo count, limit, excess count, and billing date.\nNo immediate profile changes (still active on Premium).",
     "High"),

    ("DG-05", "Downgrade",
     "Profile paused at billing cycle if over photo limit",
     "Escort downgraded, had excess photos, did not remove them. Billing period ends.",
     "1. Webhook fires subscription.updated.\n2. Check escort's is_approved status.",
     "escorts.is_approved = false. Profile disappears from search. Email sent about profile paused.",
     "DB: escorts.is_approved=false.\nSearch results: escort's profile no longer appears.",
     "High"),

    # ── CANCELLATION ─────────────────────────────────────────────────────
    ("CN-01", "Cancellation",
     "Cancel subscription — schedules end of period",
     "Escort has an active Essential monthly subscription.",
     "1. Go to My Subscriptions.\n2. Click 'Cancel Subscription'.\n3. Confirm.",
     "Subscription scheduled to cancel at end of billing period. Status shows 'Cancelling'.",
     "Stripe Dashboard: cancel_at_period_end=true.\nDB: subscriptions.status='cancelling'.\nUI: Shows 'Active until [date]' with cancel scheduled.",
     "High"),

    ("CN-02", "Cancellation",
     "Access continues until billing period end",
     "Subscription in 'cancelling' state.",
     "1. Log in as cancelling escort.\n2. Check subscription page and profile visibility.",
     "escort still has their paid tier. Profile visible in search. Features work normally.",
     "DB: escorts.subscription_tier still = paid tier.\nAPI GET /payments/subscription: status='cancelling', current_period_end in future.\nProfile appears in search.",
     "High"),

    ("CN-03", "Cancellation",
     "Tier resets to free after cancellation completes",
     "Subscription period end has passed. Stripe fires subscription.deleted webhook.",
     "1. Simulate/wait for customer.subscription.deleted webhook.\n2. Check DB + UI.",
     "Tier resets to 'free'. stripe_subscription_id cleared. Profile drops in search rank.",
     "DB: escorts.subscription_tier='free', escorts.stripe_subscription_id=null.\nDB: subscriptions.status='cancelled'.\nUI: Dashboard shows Free tier.",
     "High"),

    ("CN-04", "Cancellation",
     "Cancel when already cancelling returns error",
     "Subscription status is already 'cancelling'.",
     "1. Attempt to cancel again.",
     "API returns 400: 'Your subscription is already scheduled to cancel at the end of the billing period'.",
     "Clear error shown in UI. No duplicate Stripe calls.",
     "Medium"),

    ("CN-05", "Cancellation",
     "Reactivate subscription (undo cancellation)",
     "Subscription is in 'cancelling' state (period not yet ended).",
     "1. Go to My Subscriptions.\n2. Click 'Reactivate'.",
     "cancel_at_period_end set back to false. Status returns to 'active'. Auto-renewal resumed.",
     "Stripe Dashboard: cancel_at_period_end=false.\nDB: subscriptions.status='active'.\nUI: Shows 'Active' with renewal date.",
     "High"),

    ("CN-06", "Cancellation",
     "Cancel Blue Tick subscription",
     "Escort has active Blue Tick subscription.",
     "1. Go to Verification page or Subscription page.\n2. Cancel Blue Tick.",
     "Blue Tick subscription scheduled to cancel. blue_tick_active remains true until period end.",
     "Stripe Dashboard: Blue Tick sub cancel_at_period_end=true.\nDB: blue_tick_stripe_subscription_id sub row status='cancelling'.\nescorts.blue_tick_active still true.",
     "High"),

    ("CN-07", "Cancellation",
     "Blue Tick deactivated after subscription ends",
     "Blue Tick subscription.deleted webhook fires.",
     "1. Simulate/wait for Blue Tick customer.subscription.deleted.\n2. Check DB.",
     "blue_tick_active = false. blue_tick_stripe_subscription_id cleared. verification_level drops from 3 to 2.",
     "DB: escorts.blue_tick_active=false, blue_tick_stripe_subscription_id=null.\nescorts.verification_level=2 (not 3).",
     "High"),

    ("CN-08", "Cancellation",
     "Reactivate Blue Tick",
     "Blue Tick subscription is in 'cancelling' state.",
     "1. Click 'Reactivate Blue Tick'.",
     "cancel_at_period_end set to false. Blue Tick subscription remains active.",
     "Stripe: cancel_at_period_end=false.\nDB: blue_tick sub row status='active'.",
     "Medium"),

    # ── BLUE TICK ─────────────────────────────────────────────────────────
    ("BT-01", "Blue Tick",
     "Blue Tick checkout blocked for Premium/Elite",
     "Escort is on Premium or Elite plan.",
     "1. Go to Blue Tick section.\n2. Attempt to purchase Blue Tick.",
     "API returns 400: 'Blue Tick is included free with your Premium or Elite plan.'",
     "Error message shown. No Stripe session created.",
     "High"),

    ("BT-02", "Blue Tick",
     "Blue Tick checkout shows setup fee + monthly",
     "Escort is on Essential. No existing Blue Tick subscription.",
     "1. Go to Blue Tick section.\n2. Click 'Apply for Blue Tick'.",
     "Stripe Checkout shows two line items: £10 setup fee and £3.99/mo recurring.",
     "Stripe Dashboard: checkout session has 2 line_items.\nStripe Checkout page shows both charges clearly.",
     "High"),

    ("BT-03", "Blue Tick",
     "Blue Tick payment creates pending verification",
     "Essential escort. Blue Tick checkout completed with test card.",
     "1. Complete Blue Tick checkout.\n2. Return to site.",
     "Verification record created (level=3, status='pending'). Admin receives email notification.",
     "DB: verifications row with level=3, status='pending', escort_id correct.\nAdmin email received with link to portal.",
     "High"),

    ("BT-04", "Blue Tick",
     "Cannot buy Blue Tick twice",
     "Escort already has blue_tick_stripe_subscription_id set.",
     "1. Attempt Blue Tick checkout again.",
     "API returns 400: 'You already have an active Blue Tick subscription'.",
     "Error shown. No Stripe session created.",
     "High"),

    ("BT-05", "Blue Tick",
     "Admin approves Blue Tick",
     "Pending verification (level=3) exists in admin portal.",
     "1. Log in to admin portal.\n2. Open verification.\n3. Click Approve.",
     "blue_tick_active = true. verification_level = 3. Approval email sent to escort.",
     "DB: escorts.blue_tick_active=true, verification_level=3.\nEscort receives approval email.\nBlue Tick badge appears on public profile.",
     "High"),

    ("BT-06", "Blue Tick",
     "Admin rejects Blue Tick (refund + cancel)",
     "Pending level-3 verification exists.",
     "1. Log in to admin portal.\n2. Open verification.\n3. Enter rejection reason.\n4. Click Reject.",
     "Stripe refund issued. Blue Tick subscription cancelled immediately. Email sent to escort with reason.",
     "Stripe Dashboard: refund created for Blue Tick setup + any monthly charge.\nStripe: subscription status=cancelled.\nDB: escorts.blue_tick_active=false, blue_tick_stripe_subscription_id=null.\nEscort receives rejection email with admin notes.",
     "High"),

    # ── WEBHOOKS ──────────────────────────────────────────────────────────
    ("WH-01", "Webhooks",
     "Duplicate webhook delivery is idempotent",
     "A checkout.session.completed event has already been processed.",
     "1. Resend the same Stripe webhook event (same event ID) via Stripe CLI or dashboard.\n2. Check DB.",
     "Second delivery returns {received: true, duplicate: true}. DB unchanged (no duplicate subscription row).",
     "DB: webhook_events table has only 1 row for that event_id.\nDB: No duplicate subscription rows.\nStripe: no duplicate charges.",
     "High"),

    ("WH-02", "Webhooks",
     "checkout.session.completed creates subscription",
     "Stripe checkout completed successfully.",
     "1. Stripe fires checkout.session.completed.\n2. Check DB.",
     "Subscription record created. Escort tier updated. stripe_subscription_id saved.",
     "DB: subscriptions row with correct tier, status='active', stripe_subscription_id.\nDB: escorts.subscription_tier updated.",
     "High"),

    ("WH-03", "Webhooks",
     "subscription.updated syncs period dates",
     "Escort has active subscription. Stripe fires subscription.updated (e.g. renewal).",
     "1. Simulate subscription.updated via Stripe CLI.\n2. Check DB.",
     "current_period_start and current_period_end updated in subscriptions table.",
     "DB: subscriptions.current_period_start and current_period_end match Stripe values.",
     "Medium"),

    ("WH-04", "Webhooks",
     "subscription.updated applies pending downgrade on renewal",
     "Escort downgraded (pending_tier set). Stripe fires subscription.updated with new billing period.",
     "1. Simulate subscription.updated with period_start > current_period_start.\n2. Check DB.",
     "escorts.subscription_tier = pending_tier. subscriptions.pending_tier = null.",
     "DB: tier switched to pending_tier.\nUI: Dashboard shows downgraded tier.",
     "High"),

    ("WH-05", "Webhooks",
     "subscription.deleted resets tier to free",
     "Escort's subscription has ended/been cancelled.",
     "1. Stripe fires customer.subscription.deleted for main subscription.\n2. Check DB.",
     "escorts.subscription_tier='free'. stripe_subscription_id=null. subscriptions row status='cancelled'.",
     "DB: all fields cleared as expected.\nUI: Escort now sees Free tier.",
     "High"),

    ("WH-06", "Webhooks",
     "Webhook customer ID mismatch rejected",
     "An event arrives for a subscription ID, but the customer ID differs from our stored value.",
     "1. Manually craft/simulate a webhook with mismatched customer ID.\n2. Check server logs.",
     "Event is NOT processed. Security warning logged. DB unchanged.",
     "Server logs show '[SECURITY] ... customer mismatch'.\nDB: no changes made for that event.",
     "High"),

    # ── IDENTITY VERIFICATION ─────────────────────────────────────────────
    ("IV-01", "Identity Verification",
     "Free tier user cannot submit identity documents",
     "Escort is on Free tier (no paid subscription).",
     "1. Attempt to access /dashboard/verify.\n2. Try to upload documents.",
     "Error: 'An active paid subscription is required to submit identity verification.'",
     "API POST /api/verification/submit-identity-documents returns 403/400.\nUI: Upload button not shown for free tier.",
     "High"),

    ("IV-02", "Identity Verification",
     "Submit identity documents (happy path)",
     "Escort is on Essential or higher. No pending verification.",
     "1. Go to /dashboard/verify.\n2. Upload ID document (JPEG).\n3. Upload selfie.\n4. Submit.",
     "Verification record created with status='pending'. Admin notification email sent.",
     "DB: verifications row with level=2, status='pending', id_document_url and selfie_url set.\nAdmin email received.",
     "High"),

    ("IV-03", "Identity Verification",
     "Cannot subscribe while identity verification is pending",
     "Escort has pending level-2 verification.",
     "1. Attempt to start a new subscription checkout.",
     "API returns 409: 'Your current application is being reviewed. Please wait before subscribing to another tier.'",
     "POST /payments/checkout returns 409.\nUI: Subscribe button shows 'Pending Review' state.",
     "High"),

    ("IV-04", "Identity Verification",
     "Admin approves identity verification",
     "Pending level-2 verification exists.",
     "1. Admin logs in to portal.\n2. Reviews documents.\n3. Clicks Approve.",
     "verification_level = 2. Approval email sent to escort.",
     "DB: escorts.verification_level=2.\nDB: verifications.status='approved'.\nEscort receives approval email.",
     "High"),

    ("IV-05", "Identity Verification",
     "Admin rejects identity verification (refund + cancel)",
     "Pending level-2 verification. Escort has an active paid subscription.",
     "1. Admin logs in.\n2. Enters rejection reason.\n3. Clicks Reject.",
     "Stripe refund issued. Subscription cancelled. Tier resets to free. Email sent with reason.",
     "Stripe Dashboard: refund created.\nDB: escorts.subscription_tier='free'.\nDB: verifications.status='rejected', admin_notes set.\nEscort receives rejection email.",
     "High"),

    # ── REFERRAL PROGRAMME ────────────────────────────────────────────────
    ("RF-01", "Referral",
     "Apply referral code at checkout",
     "Referrer has a referral_code. New user knows this code.",
     "1. Go to checkout.\n2. Enter referral code in referral field.\n3. Complete payment.",
     "50% discount applied for 3 months. Coupon attached to Stripe subscription.",
     "Stripe: subscription has 50% off × 3 months coupon.\nDB: escort.referred_by_code = referrer's code.",
     "High"),

    ("RF-02", "Referral",
     "Cannot use own referral code",
     "Escort has a referral_code.",
     "1. Enter your own referral code at checkout.",
     "API returns 400: 'You cannot use your own referral code'.",
     "Error message shown. No session created.",
     "High"),

    ("RF-03", "Referral",
     "Referrer receives 1 month credit on first payment",
     "Escort A referred Escort B. Escort B makes their first payment.",
     "1. Stripe fires invoice.paid for Escort B's first subscription invoice.\n2. Check Stripe for Escort A.",
     "Escort A receives a Stripe customer balance credit equal to 1 month of their plan.",
     "Stripe Dashboard: Escort A's customer balance shows negative credit (= 1 month's amount).\nDB: escort B's referral_reward_claimed=true.",
     "High"),

    ("RF-04", "Referral",
     "Referral reward only granted once",
     "Escort B's first invoice already triggered a reward for Escort A.",
     "1. Escort B renews (second invoice.paid fires).\n2. Check Stripe for Escort A.",
     "No additional credit added. Reward only granted once.",
     "DB: escort B's referral_reward_claimed=true (unchanged).\nStripe: no new balance transaction for Escort A.",
     "Medium"),

    # ── UI / DASHBOARD ────────────────────────────────────────────────────
    ("UI-01", "UI / Dashboard",
     "Subscription page shows current plan and status",
     "Logged in escort with active Essential subscription.",
     "1. Go to Subscription page.",
     "Page shows: current tier (Essential), status (Active), next billing date, cancel button.",
     "GET /api/payments/subscription returns correct data.\nUI renders tier, status, and period_end correctly.",
     "High"),

    ("UI-02", "UI / Dashboard",
     "Photo limit enforced per tier",
     "Escort on Essential (limit: 8). Already has 8 photos.",
     "1. Go to Photos page.\n2. Attempt to upload a 9th photo.",
     "Upload rejected with message about photo limit for current tier.",
     "API POST /upload/photo returns 400 about limit.\nUI: Upload button disabled or shows limit message.",
     "High"),

    ("UI-03", "UI / Dashboard",
     "Invoice history populates after payment",
     "Escort has completed at least one payment.",
     "1. Go to My Subscriptions / Invoices.",
     "Invoice list shows at least one entry with amount, date, and PDF link.",
     "GET /api/payments/invoices returns non-empty list.\nPDF link opens invoice in browser.",
     "High"),

    ("UI-04", "UI / Dashboard",
     "Free tier CTA drives to subscription page",
     "Escort on Free tier, logged in.",
     "1. Go to Dashboard.",
     "CTA / upsell section is visible encouraging upgrade. 'Subscribe' or 'Upgrade' button visible.",
     "UI renders upgrade prompt for free tier.\nSubscription page opens correctly from CTA.",
     "Medium"),

    ("UI-05", "UI / Dashboard",
     "Profile search ranking: Elite > Premium > Essential > Free",
     "4 test escorts: one of each tier, all with complete profiles and active.",
     "1. Go to /escorts (search page) with no filters.",
     "Elite escort appears first, then Premium, then Essential, then Free.",
     "Visual order on search page matches tier ordering.\nAPI GET /escorts returns results in correct order.",
     "High"),

    ("UI-06", "UI / Dashboard",
     "Available Now toggle affects profile ordering",
     "Two Essential escorts: one with available_now=true, one false.",
     "1. Search /escorts.",
     "The escort with available_now=true appears higher in results among the same tier.",
     "Visual order: available_now=true escorts ranked above available_now=false within same tier.",
     "Medium"),

    # ── EDGE CASES ────────────────────────────────────────────────────────
    ("EC-01", "Edge Cases",
     "Checkout without email verification blocked",
     "Escort registered but has NOT verified their email.",
     "1. Attempt to reach checkout (POST /payments/checkout).",
     "API returns 403: email not verified.",
     "get_current_verified_escort dependency blocks unverified escorts.\nUI: prompt to verify email.",
     "High"),

    ("EC-02", "Edge Cases",
     "Checkout with Stripe keys not configured returns 503",
     "STRIPE_SECRET_KEY is blank (test config only).",
     "1. POST /payments/checkout.",
     "API returns 503: 'Payments not configured'.",
     "503 response with clear message.",
     "Low"),

    ("EC-03", "Edge Cases",
     "Upgrade with no existing subscription returns error",
     "Escort has no stripe_subscription_id.",
     "1. POST /payments/upgrade-tier.",
     "API returns 400: 'No active subscription found. Use the checkout flow to subscribe.'",
     "Clear error message. Escort directed to checkout flow.",
     "Medium"),

    ("EC-04", "Edge Cases",
     "Cancel with no active subscription returns error",
     "Escort has subscription_tier='free' and no stripe_subscription_id.",
     "1. POST /payments/cancel.",
     "API returns 400: 'No active subscription found'.",
     "Error shown in UI.",
     "Medium"),

    ("EC-05", "Edge Cases",
     "Annual Blue Tick billing anchor set to 1st of next month",
     "Escort purchases Blue Tick mid-month.",
     "1. Complete Blue Tick checkout.\n2. Check Stripe.",
     "Blue Tick subscription anchored to 1st of next month, same as main subscription.",
     "Stripe: blue_tick sub billing_cycle_anchor = 1st of next month.",
     "Low"),

    ("EC-06", "Edge Cases",
     "Invalid tier in checkout request rejected",
     "Any authenticated escort.",
     "1. POST /payments/checkout with tier='superduper'.",
     "API returns 400: 'Invalid tier. Choose: essential, premium, or elite'.",
     "400 response with clear error.",
     "Low"),
]


def create_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws_tests: Worksheet = wb.active
    ws_tests.title = "Test Cases"

    _build_test_sheet(ws_tests)
    _build_summary_sheet(wb)

    return wb


def _cell_style(ws, row, col, value, bg=None, fg="000000", bold=False,
                wrap=True, align="left", border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(color=fg, bold=bold, size=10,
                     name="Calibri")
    cell.alignment = Alignment(wrap_text=wrap, vertical="top",
                                horizontal=align)
    if border:
        cell.border = THIN_BORDER
    return cell


def _build_test_sheet(ws: Worksheet):
    # Title row
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Bluechips London — Payment & Subscription Test Plan"
    title_cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    title_cell.font = Font(color=HEADER_FG, bold=True, size=14, name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Sub-header row
    ws.merge_cells("A2:J2")
    sub = ws["A2"]
    sub.value = "Use Stripe test cards: 4242 4242 4242 4242 (success) · 4000 0000 0000 0002 (decline) · 4000 0025 0000 3155 (3DS)"
    sub.fill = PatternFill("solid", fgColor="2D2D44")
    sub.font = Font(color="AAAAAA", italic=True, size=9, name="Calibri")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Column headers (row 3)
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.font = Font(color=HEADER_FG, bold=True, size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[3].height = 22

    # Freeze panes: freeze title + header rows and ID/Category/Name cols
    ws.freeze_panes = "D4"

    # Write test rows
    current_category = None
    row = 4
    for test in TESTS:
        tid, cat, name, precond, steps, expected, validation, priority = test

        # Category separator row
        if cat != current_category:
            current_category = cat
            ws.merge_cells(f"A{row}:J{row}")
            cat_cell = ws[f"A{row}"]
            cat_cell.value = f"  {cat.upper()}"
            cat_cell.fill = PatternFill("solid", fgColor=CAT_BG)
            cat_cell.font = Font(color=CAT_FG, bold=True, size=10, name="Calibri")
            cat_cell.alignment = Alignment(vertical="center")
            cat_cell.border = THIN_BORDER
            ws.row_dimensions[row].height = 18
            row += 1

        # Alternate row background
        row_bg = ALT_ROW if row % 2 == 0 else WHITE
        pri_bg = PRIORITY_COLOURS.get(priority, WHITE)

        values = [tid, cat, name, precond, steps, expected, validation, priority,
                  "Not Run", ""]
        for col_idx, val in enumerate(values, start=1):
            bg = pri_bg if col_idx == 8 else (STATUS_BG if col_idx == 9 else row_bg)
            fg_color = "000000"
            if col_idx == 1:
                fg_color = "555555"
                bold = True
            else:
                bold = False
            _cell_style(ws, row, col_idx, val, bg=bg, fg=fg_color, bold=bold)

        # Taller rows for multi-line content
        ws.row_dimensions[row].height = max(
            60,
            max(len(str(v)) for v in [steps, expected, validation]) // 3
        )
        row += 1

    # Status dropdown (col 9)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"Not Run,Pass,Fail,Blocked,Skip"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.sqref = f"I4:I{row}"

    # Priority dropdown (col 8)
    dv_pri = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(dv_pri)
    dv_pri.sqref = f"H4:H{row}"


def _build_summary_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Summary")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12

    # Title
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "Test Plan Summary"
    t.fill = PatternFill("solid", fgColor=HEADER_BG)
    t.font = Font(color=HEADER_FG, bold=True, size=13, name="Calibri")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers
    headers = ["Category", "Total", "High", "Medium", "Low", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.font = Font(color=HEADER_FG, bold=True, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    ws.row_dimensions[2].height = 18

    # Collect counts
    from collections import Counter, defaultdict
    cat_counts = defaultdict(lambda: {"total": 0, "High": 0, "Medium": 0, "Low": 0})
    for t in TESTS:
        tid, cat, name, *_, priority = t
        cat_counts[cat]["total"] += 1
        cat_counts[cat][priority] += 1

    row = 3
    totals = {"total": 0, "High": 0, "Medium": 0, "Low": 0}
    for cat, counts in cat_counts.items():
        bg = ALT_ROW if row % 2 == 0 else WHITE
        ws.cell(row=row, column=1, value=cat).border = THIN_BORDER
        ws.cell(row=row, column=2, value=counts["total"]).border = THIN_BORDER
        ws.cell(row=row, column=3, value=counts["High"]).border = THIN_BORDER
        ws.cell(row=row, column=4, value=counts["Medium"]).border = THIN_BORDER
        ws.cell(row=row, column=5, value=counts["Low"]).border = THIN_BORDER
        ws.cell(row=row, column=6, value="").border = THIN_BORDER
        for col in range(1, 7):
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(size=10, name="Calibri")
            c.alignment = Alignment(horizontal="center" if col > 1 else "left")
        for k in totals:
            totals[k] += counts[k]
        row += 1

    # Totals row
    for col, val in enumerate([
        "TOTAL",
        totals["total"],
        totals["High"],
        totals["Medium"],
        totals["Low"],
        ""
    ], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=CAT_BG)
        c.font = Font(color=CAT_FG, bold=True, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center" if col > 1 else "left")
        c.border = THIN_BORDER

    row += 2
    ws.cell(row=row, column=1, value="Legend").font = Font(bold=True, size=10)
    row += 1
    legend = [
        (HIGH_BG, "High priority — must pass before going live"),
        (MED_BG,  "Medium priority — should pass for full confidence"),
        (LOW_BG,  "Low priority — edge case / nice to verify"),
    ]
    for bg, label in legend:
        c = ws.cell(row=row, column=1, value="  " + label)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(size=10, name="Calibri")
        row += 1


if __name__ == "__main__":
    wb = create_workbook()
    wb.save(OUTPUT)
    print(f"Test plan written to: {OUTPUT}")
    total = len(TESTS)
    high = sum(1 for t in TESTS if t[-1] == "High")
    print(f"Total tests: {total} ({high} High priority)")
